# internal
from src.ui.component import BaseDialog, FireButton, CheckList
# external
from openpyxl import load_workbook
# pyqt
from PyQt5.QtCore import Qt
from PyQt5.Qt import QListView, QAbstractItemView
from PyQt5.QtWidgets import QHBoxLayout, QFileDialog


class ExcelResponse(BaseDialog):
    """Excel Responsible"""
    def craftDialog(self):
        # self modification
        self.setWindowTitle('Excel Titles')
        self.setFixedWidth(320)
        # excel table
        self.listExcel = QListView()
        # done button
        self.btnDoneLayout = QHBoxLayout()
        self.btnDoneLayout.setAlignment(Qt.AlignHCenter)
        self.btnDone = FireButton('Done')
        # attach
        self.generalLayout.addWidget(self.listExcel)
        self.btnDoneLayout.addWidget(self.btnDone)
        self.generalLayout.addLayout(self.btnDoneLayout)

    def openExcel(self):
        # empty list
        self.titleList = list()
        path = QFileDialog.getOpenFileName(
            self, 'Open file', '',
            'Excel files (*.xlsx *.xlsm *.xltx *.xltm)')[0]
        if path:
            wb = load_workbook(path)
            sheet = wb.active
            for cell in sheet.iter_cols(max_row=1, values_only=True):
                index = cell[0]
                if index:
                    strip_index = str(index).strip()
                    self.titleList.append(strip_index)

        self.modelExcel = CheckList(self.titleList)
        self.listExcel.setModel(self.modelExcel)
        self.listExcel.setEditTriggers(QAbstractItemView.NoEditTriggers)

        if path and self.titleList:
            return True

    def craftStyles(self):
        self.setStyleSheet("""
            ExcelResponse {
                background: #fcfcfc;
            }
        """)

