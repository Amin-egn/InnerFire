# internal
from src.ui.component import (BaseWidget, BaseDialog, FireButton, TableModel,
                              ListModel, CheckList, TableView)
# external
from openpyxl import load_workbook
# pyqt
from PyQt5.QtCore import (Qt)
from PyQt5.QtGui import QLinearGradient
from PyQt5.QtWidgets import (QMainWindow, QWidget, QPushButton, QHBoxLayout, QFileDialog,
                             QLineEdit, QFormLayout, QLabel, QTableView, QVBoxLayout, QListView,
                             QAbstractItemView)


class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        # bootstrap
        self._bootstrap()
        # stylesheet
        self._styleSheet()

    def _bootstrap(self):
        self.setWindowTitle('Inner Fire')
        self.setMinimumSize(640, 480)
        # central widget
        self._mainWidget = QWidget(self)
        self.setCentralWidget(self._mainWidget)
        # general layout
        self.generalLayout = QVBoxLayout()
        # set align top
        self.generalLayout.setAlignment(Qt.AlignTop)
        # front widget
        self.frontWidget = FrontWidget()
        # excel response
        self.excelResponse = ExcelResponse()
        # database response
        self.dbResponse = DbResponse()
        # database table
        self.dbTable = DbTables()
        # database titles
        self.dbTableTitles = DbTableTitles()
        # attach
        self._mainWidget.setLayout(self.generalLayout)
        self.generalLayout.addWidget(self.frontWidget)

    def _styleSheet(self):
        self.setStyleSheet("""
            MainWindow {
                background: #fcfcfc;
            }
        """)

class FrontWidget(BaseWidget):
    """Front Widget"""
    def craftWidget(self):
        self.actionButtons()
        self.tableFrame()
        self.castButton()

    def actionButtons(self):
        # button layout
        self.buttonsLayout = QHBoxLayout()
        self.buttonsLayout.setContentsMargins(0, 10, 0, 0)
        self.buttonsLayout.setAlignment(Qt.AlignHCenter)
        # excel button
        self.btnExcel = FireButton('Read Excel', icon='./src/ui/resources/spreadsheet.png')
        # data base button
        self.btnDb = FireButton('Query Data-Base', icon='./src/ui/resources/database.png')
        # attach
        self.buttonsLayout.addWidget(self.btnExcel)
        self.buttonsLayout.addWidget(self.btnDb)
        self.generalLayout.addLayout(self.buttonsLayout)

    def listFrame(self):
        pass

    def tableFrame(self):
        # frame widget
        self.tableWidget = QTableView()
        self.tableWidget.setObjectName('Table')
        # frame layout
        self.tableWidget.setMinimumSize(600, 360)
        self.tableLayout = QVBoxLayout()
        # attach
        self.tableWidget.setLayout(self.tableLayout)
        self.generalLayout.addWidget(self.tableWidget)

    def castButton(self):
        # cast layout
        self.castLayout = QHBoxLayout()
        self.castLayout.setContentsMargins(0, 10, 0, 5)
        # cast label
        self.lblCast = QLabel('You should select excel and sql table that you want to cast =)')
        # cast button
        self.btnCast = FireButton('Cast!')
        # attach
        self.castLayout.addWidget(self.lblCast)
        self.castLayout.addWidget(self.btnCast)
        self.generalLayout.addLayout(self.castLayout)

    def craftStyle(self):
        self.setStyleSheet("""
            #Table {
                border: 2px dot-dash #3d9049;
            }
        """)


class ExcelResponse(BaseDialog):
    """Excel Responsible"""
    def craftDialog(self):
        # self modification
        self.setWindowTitle('Excel Titles')
        self.setFixedWidth(320)
        # excel table
        self.listExcel = QListView()
        # done button
        self.btnDoneLayout = QHBoxLayout()
        self.btnDoneLayout.setAlignment(Qt.AlignHCenter)
        self.btnDone = FireButton('Done')
        # attach
        self.generalLayout.addWidget(self.listExcel)
        self.btnDoneLayout.addWidget(self.btnDone)
        self.generalLayout.addLayout(self.btnDoneLayout)

    def openExcel(self):
        # empty list
        self.titleList = list()
        path = QFileDialog.getOpenFileName(
            self, 'Open file', '',
            'Excel files (*.xlsx *.xlsm *.xltx *.xltm)')[0]
        if path:
            wb = load_workbook(path)
            sheet = wb.active
            for cell in sheet.iter_cols(max_row=1, values_only=True):
                index = cell[0]
                if index:
                    strip_index = str(index).strip()
                    self.titleList.append(strip_index)

        self.modelExcel = CheckList(self.titleList)
        self.listExcel.setModel(self.modelExcel)
        self.listExcel.setEditTriggers(QAbstractItemView.NoEditTriggers)

        if path and self.titleList:
            return True

    def craftStyles(self):
        self.setStyleSheet("""
            ExcelResponse {
                background: #fcfcfc;
            }
        """)


class DbResponse(BaseDialog):
    """Database Responsible"""
    def craftDialog(self):
        # self modification
        self.setFixedWidth(380)
        self.setWindowTitle('Data-base Information')
        # form layout
        self.dataBaseLayout = QFormLayout()
        # inputs
        self.serverInput = QLineEdit(text='.\\')
        self.usernameInput = QLineEdit(text='sa')
        self.passwordInput = QLineEdit(echoMode=QLineEdit.EchoMode.Password)
        self.dbnameInput = QLineEdit()
        # add widgets
        self.dataBaseLayout.addRow('Server:', self.serverInput)
        self.dataBaseLayout.addRow('Username:', self.usernameInput)
        self.dataBaseLayout.addRow('Password:', self.passwordInput)
        self.dataBaseLayout.addRow('Database name:', self.dbnameInput)
        # button layout
        self.btnLayout = QHBoxLayout()
        self.btnLayout.setContentsMargins(0, 10, 0, 0)
        self.btnConnect = FireButton('Connect')
        # attach
        self.btnLayout.addWidget(self.btnConnect)
        self.generalLayout.addLayout(self.dataBaseLayout)
        self.generalLayout.addLayout(self.btnLayout)

    def connList(self):
        self.inputList = [
            self.serverInput.text().strip(),
            self.usernameInput.text().strip(),
            self.passwordInput.text().strip(),
            self.dbnameInput.text().strip()
        ]
        return self.inputList

    def connectSignals(self):
        self.btnConnect.clicked.connect(self.connList)

    def craftStyles(self):
        self.setStyleSheet("""
            DbResponse {
                background: #fcfcfc;
            }
        """)


class DbTables(BaseDialog):
    """Database Tables"""
    def craftDialog(self):
        # self modification
        self.setWindowTitle('Data-base Tables')
        # tables list
        self.listTables = QListView()
        self.listTables.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # done button
        self.btnDoneLayout = QHBoxLayout()
        self.btnDoneLayout.setAlignment(Qt.AlignHCenter)
        self.btnDone = FireButton('Done')
        # attach
        self.btnDoneLayout.addWidget(self.btnDone)
        self.generalLayout.addWidget(self.listTables)
        self.generalLayout.addLayout(self.btnDoneLayout)

    def craftStyles(self):
        self.setStyleSheet("""
            DbTables {
                background: #fcfcfc;
            }
        """)


class DbTableTitles(BaseDialog):
    """Database Table Titles"""
    def craftDialog(self):
        # self modification
        self.setWindowTitle('Table Titles')
        # tables list
        self.listTitles = QListView()
        # done button
        self.btnDoneLayout = QHBoxLayout()
        self.btnDoneLayout.setAlignment(Qt.AlignHCenter)
        self.btnDone = FireButton('Done')
        # attach
        self.btnDoneLayout.addWidget(self.btnDone)
        self.generalLayout.addWidget(self.listTitles)
        self.generalLayout.addLayout(self.btnDoneLayout)

    def craftStyles(self):
        self.setStyleSheet("""
            DbTableTitles {
                background: #fcfcfc;
            }
        """)
