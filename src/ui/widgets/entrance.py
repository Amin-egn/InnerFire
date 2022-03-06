# internal
from .base import BaseWidget
from .excel import ExcelResponse
from .database import DbResponse
from src.ui.component import FireButton, LockedTableView, PresentationTableWidget
# external
from openpyxl import load_workbook
# pyqt
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QHBoxLayout, QLabel, QFileDialog, QFrame


class Entrance(BaseWidget):
    """Entrance"""
    def _initialize(self):
        # Excel collector list
        self.excelDataCollectorList = list()
        # table collector list
        self.dbDataCollectorDict = dict()

    def _craftWidget(self):
        self._actionButtons()
        # self._tableWidget()
        self._modelsFrame()
        self._nextStage()
        # Excel
        self.excelResponse = ExcelResponse(self)
        # database
        # - response
        self.dbResponse = DbResponse(self)

    def _actionButtons(self):
        # button layout
        self.buttonsLayout = QHBoxLayout()
        self.buttonsLayout.setContentsMargins(0, 10, 0, 0)
        self.buttonsLayout.setAlignment(Qt.AlignHCenter)
        # excel button
        self.btnExcel = FireButton('Read Excel', icon=':/icons/excel')
        # data base button
        self.btnDb = FireButton('Query Data-Base', icon=':/icons/database')
        # attach
        self.buttonsLayout.addWidget(self.btnExcel)
        self.buttonsLayout.addWidget(self.btnDb)
        # - general
        self.generalLayout.addLayout(self.buttonsLayout)

    def _modelsFrame(self):
        # frame
        self.modelsFrame = QFrame(self)
        self.modelsFrame.setMinimumSize(580, 360)
        self.modelsFrame.setObjectName('Table')
        # layout
        self.modelsLayout = QHBoxLayout()
        self.modelsFrame.setLayout(self.modelsLayout)
        # views
        # - excel
        self.excelTableView = LockedTableView(None)
        self.excelTableView.hide()
        # - database
        self.dbTableWidget = PresentationTableWidget()
        self.dbTableWidget.hide()
        # attach
        # - main
        self.modelsLayout.addWidget(self.excelTableView)
        self.modelsLayout.addWidget(self.dbTableWidget)
        # - general
        self.generalLayout.addWidget(self.modelsFrame)

    def _nextStage(self):
        # cast layout
        self.nextStageLayout = QHBoxLayout()
        self.nextStageLayout.setContentsMargins(0, 10, 0, 5)
        # cast label
        self.lblNextStage = QLabel('You should select excel and sql table that you want to cast =)')
        # cast button
        self.btnNextStage = FireButton('Next Stage!')
        self.btnNextStage.setDisabled(True)
        # attach
        self.nextStageLayout.addWidget(self.lblNextStage)
        self.nextStageLayout.addWidget(self.btnNextStage)
        # - general
        self.generalLayout.addLayout(self.nextStageLayout)

    def _openExcel(self):
        # empty list
        self.titleList = list()
        path = QFileDialog.getOpenFileName(
            self, 'Open file', '',
            'Excel files (*.xlsx *.xlsm *.xltx *.xltm)')[0]

        if path:
            wb = load_workbook(path, data_only=True)
            self.sheet = wb.active
            for cell in self.sheet.iter_cols(max_row=1, values_only=True):
                index = cell[0]
                if index:
                    strip_index = str(index).strip()
                    self.titleList.append(strip_index)

        if self.titleList:
            self.excelDataCollectorList.clear()
            self.excelResponse.excelTitleList.clear()
            self.excelTableView.setModel(None)
            self.excelResponse.selectedTitleTableModel.clearRecords()
            # excel
            self.excelResponse.excelListModel.setItems(self.titleList)
            self.excelResponse.exec()

    def _database(self):
        self.dbResponse.exec()

    def createExcelView(self, model):
        if model:
            self.excelTableView.setModel(model)
            self.excelTableView.show()

        else:
            self.excelDataCollectorList.clear()
            self.excelTableView.setModel(None)

    def createDbTable(self, headers, records):
        if headers and records:
            self.dbTableWidget.setTitles(headers)
            self.dbTableWidget.setRowRecords(records)
            self.dbTableWidget.show()

        else:
            self.dbTableWidget.clear()
            self.dbDataCollectorDict.clear()

    def unlockNextStage(self):
        if self.excelDataCollectorList and self.dbDataCollectorDict:
            self.btnNextStage.setEnabled(True)
        else:
            self.btnNextStage.setDisabled(True)
            if not self.excelDataCollectorList:
                self.excelTableView.hide()
            if not self.dbDataCollectorDict:
                self.dbTableWidget.hide()

    # noinspection PyUnresolvedReferences
    def _connectSignals(self):
        self.btnExcel.clicked.connect(self._openExcel)
        self.btnDb.clicked.connect(self._database)

    def _craftStyle(self):
        self.setStyleSheet("""
            MainWindow {
                background: #fcfcfc;
            }
            #Table {
                border: 2px dot-dash #33892a;
            }
        """)
